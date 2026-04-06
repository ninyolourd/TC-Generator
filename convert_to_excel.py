import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def parse_md(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    rows = []
    us_blocks = re.split(r'\n(?=## User Story)', content)

    for block in us_blocks:
        us_match = re.match(r'## (User Story \d+): (.+)', block)
        if not us_match:
            continue
        us_num = us_match.group(1)
        us_title = us_match.group(2).strip()

        # Grab the full "As a..." description line if present
        us_desc_match = re.search(r'Module:.+\n(As .+)', block)
        if us_desc_match:
            us_full = f'{us_num}: {us_title}\n{us_desc_match.group(1).strip()}'
        else:
            us_full = f'{us_num}: {us_title}'

        module_match = re.search(r'Module:\s*(.+)', block)
        module = module_match.group(1).strip() if module_match else ''

        tc_blocks = re.split(r'\n(?=### )', block)
        for tc_block in tc_blocks[1:]:
            tc_match = re.match(r'### (TC-[\w-]+): (.+)', tc_block)
            if not tc_match:
                continue
            tc_id = tc_match.group(1).strip()
            tc_title = tc_match.group(2).strip()

            pre_match = re.search(r'(?:Preconditions?|Pre-?conditions?)[:\*\s]*\n(.*?)(?=\n\s*\n|\nSteps?[:\*\s]*\n|\nExpected Results?[:\*\s]*\n|---|\Z)', tc_block, re.DOTALL | re.IGNORECASE)
            preconditions = ''
            if pre_match:
                lines = [l.strip().lstrip('-').strip() for l in pre_match.group(1).strip().splitlines() if l.strip()]
                preconditions = '\n'.join(lines)

            steps_match = re.search(r'Steps?\s*[:\n](.*?)(?=\n(?:Expected Results?|---|\Z))', tc_block, re.DOTALL | re.IGNORECASE)
            steps = ''
            if steps_match:
                lines = [l.strip() for l in steps_match.group(1).strip().splitlines() if l.strip()]
                steps = '\n'.join(lines)

            exp_match = re.search(r'Expected Results?\s*[:\n](.*?)(?=\n---|\Z)', tc_block, re.DOTALL | re.IGNORECASE)
            expected_lines = []
            status = 'Not Started'
            if exp_match:
                raw = exp_match.group(1).strip()
                status_match = re.search(r'\|\s*Status:\s*(.+)', raw)
                if status_match:
                    status = status_match.group(1).strip()
                    raw = re.sub(r'\s*\|\s*Status:.+', '', raw).strip()
                expected_lines = [l.strip() for l in raw.splitlines() if l.strip()]

            if not expected_lines:
                expected_lines = ['']

            # One row per expected result line
            for idx, exp_line in enumerate(expected_lines):
                rows.append({
                    'User Story': us_full,
                    'Module': module,
                    'Test Case ID': tc_id,
                    'Test Case Title': tc_title,
                    'Preconditions': preconditions if idx == 0 else '',
                    'Steps': steps if idx == 0 else '',
                    'Expected Results': exp_line,
                    'Status': status,
                    '_tc_first_row': idx == 0,
                    '_tc_key': f'{us_num}|{tc_id}',
                })

    return rows

def create_excel(rows, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Test Cases'

    headers = ['User Story', 'Module', 'Test Case ID', 'Test Case Title', 'Preconditions', 'Steps', 'Expected Results', 'Status']

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap_top = Alignment(vertical='top', wrap_text=True)

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    ws.row_dimensions[1].height = 30

    fill_even = PatternFill(start_color='EBF3FB', end_color='EBF3FB', fill_type='solid')
    fill_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    status_colors = {
        'Not Started': 'D9D9D9',
        'Pass': 'C6EFCE',
        'Fail': 'FFC7CE',
        'In Progress': 'FFEB9C',
    }

    # Group rows by tc_key and us for merging
    tc_excel_rows = {}   # tc_key -> list of excel row numbers
    us_excel_rows = {}   # us -> list of excel row numbers

    # Track alternating color per TC group
    tc_color_index = {}
    tc_counter = 0

    for i, row in enumerate(rows, 2):
        tc_key = row['_tc_key']
        us = row['User Story']

        if tc_key not in tc_excel_rows:
            tc_excel_rows[tc_key] = []
            tc_color_index[tc_key] = tc_counter % 2
            tc_counter += 1
        tc_excel_rows[tc_key].append(i)

        if us not in us_excel_rows:
            us_excel_rows[us] = []
        us_excel_rows[us].append(i)

        fill = fill_even if tc_color_index[tc_key] == 0 else fill_odd

        for col, key in enumerate(headers, 1):
            val = row[key]
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = border
            cell.fill = fill
            cell.alignment = wrap_top

            if key == 'Status' and val:
                color = status_colors.get(val, 'D9D9D9')
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

        ws.row_dimensions[i].height = 20

    # Merge cells for columns that repeat within a TC group (NOT Status - each row has its own)
    merge_cols = {
        'User Story': 1,
        'Module': 2,
        'Test Case ID': 3,
        'Test Case Title': 4,
        'Preconditions': 5,
        'Steps': 6,
    }

    for tc_key, row_nums in tc_excel_rows.items():
        if len(row_nums) > 1:
            for col in merge_cols.values():
                ws.merge_cells(start_row=row_nums[0], start_column=col, end_row=row_nums[-1], end_column=col)
                cell = ws.cell(row=row_nums[0], column=col)
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    col_widths = [38, 20, 20, 35, 30, 40, 45, 15]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    wb.save(output_path)
    total_tcs = len(tc_excel_rows)
    print(f'Saved: {output_path} ({total_tcs} test cases, {len(rows)} rows)')

def create_excel_by_module(all_rows, output_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Get ordered unique modules
    seen = []
    for r in all_rows:
        if r['Module'] not in seen:
            seen.append(r['Module'])

    for module in seen:
        rows = [r for r in all_rows if r['Module'] == module]
        # Sanitize sheet name (max 31 chars, no special chars)
        sheet_name = re.sub(r'[\\/*?:\[\]]', '', module)[:31]
        ws = wb.create_sheet(title=sheet_name)

        headers = ['User Story', 'Module', 'Test Case ID', 'Test Case Title', 'Preconditions', 'Steps', 'Expected Results', 'Status']

        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wrap_top = Alignment(vertical='top', wrap_text=True)
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        ws.row_dimensions[1].height = 30

        fill_even = PatternFill(start_color='EBF3FB', end_color='EBF3FB', fill_type='solid')
        fill_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        status_colors = {
            'Not Started': 'D9D9D9',
            'Pass': 'C6EFCE',
            'Fail': 'FFC7CE',
            'In Progress': 'FFEB9C',
        }

        tc_excel_rows = {}
        tc_color_index = {}
        tc_counter = 0

        for i, row in enumerate(rows, 2):
            tc_key = row['_tc_key']
            if tc_key not in tc_excel_rows:
                tc_excel_rows[tc_key] = []
                tc_color_index[tc_key] = tc_counter % 2
                tc_counter += 1
            tc_excel_rows[tc_key].append(i)

            fill = fill_even if tc_color_index[tc_key] == 0 else fill_odd

            for col, key in enumerate(headers, 1):
                val = row[key]
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = border
                cell.fill = fill
                cell.alignment = wrap_top

                if key == 'Status' and val:
                    color = status_colors.get(val, 'D9D9D9')
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

            ws.row_dimensions[i].height = 20

        merge_cols = [1, 2, 3, 4, 5, 6]
        for tc_key, row_nums in tc_excel_rows.items():
            if len(row_nums) > 1:
                for col in merge_cols:
                    ws.merge_cells(start_row=row_nums[0], start_column=col, end_row=row_nums[-1], end_column=col)
                    cell = ws.cell(row=row_nums[0], column=col)
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

        col_widths = [38, 20, 20, 35, 30, 40, 45, 15]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

        print(f'  Sheet "{sheet_name}": {len(tc_excel_rows)} test cases, {len(rows)} rows')

    wb.save(output_path)
    print(f'Saved: {output_path}')

def normalize_title(title):
    """Ensure title starts with 'Verify'. Rewrites system-requirement style titles."""
    if title.lower().startswith('verify'):
        return title

    # Map common patterns to Verify-style titles
    patterns = [
        (r'The system shall allow the teller to input a policy number.*[Vv]alid [Ii]nput', 'Verify system accepts valid policy number input'),
        (r'The system shall allow the teller to input a policy number.*[Ii]nvalid [Ii]nput', 'Verify system rejects invalid policy number input'),
        (r'The system shall allow the teller to input a policy number', 'Verify system allows teller to input a policy number'),
        (r'The system shall display.*policy information.*[Vv]alid', 'Verify system displays policy information for valid policy number'),
        (r'The system shall display.*policy information', 'Verify system displays correct policy information'),
        (r'Upon entry of a policy number.*not found in Ingenium', 'Verify system searches SUKI when policy not found in Ingenium'),
        (r'Upon entry of a policy number.*not found in SUKI', 'Verify system searches AUS when policy not found in SUKI'),
        (r'Upon entry of a policy number.*not found in AUS', 'Verify system displays error when policy not found in any source'),
        (r'Upon entry of a policy number.*[Zz]ero [Aa]mount', 'Verify system validates zero amount on policy number entry'),
        (r'Upon entry of a policy number.*[Bb]efore [Dd]ue [Dd]ate', 'Verify system processes payment before due date'),
        (r'Upon entry of a policy number.*[Oo]n [Dd]ue [Dd]ate', 'Verify system processes payment on due date'),
        (r'Upon entry of a policy number.*[Ww]ithin.*grace', 'Verify system processes payment within grace period'),
        (r'Upon entry of a policy number.*[Bb]eyond.*grace', 'Verify system handles payment beyond grace period'),
        (r'Upon entry of a policy number', 'Verify system retrieves policy details on policy number entry'),
        (r'The system shall connect to GPAS', 'Verify system connects to GPAS to retrieve policy information'),
        (r'The system shall allow the user to select.*[Aa]gent [Tt]ran[Cc]ode', 'Verify system allows selection of applicable Agent TranCode'),
        (r'The system shall allow the user to select.*[Tt]ran[Cc]ode.*filter', 'Verify system filters TranCode options based on policy type'),
        (r'The system shall allow the user to select.*[Tt]ran[Cc]ode', 'Verify system allows user to select applicable TranCode'),
        (r'The system shall allow the user to enter.*[Tt]ransaction [Aa]mount', 'Verify system allows user to enter transaction amount'),
        (r'The system shall allow editing of the premium amount', 'Verify system allows editing of the premium amount'),
        (r'The system shall allow.*[Tt]op.?[Uu]p', 'Verify system allows top-up transaction'),
        (r'The system shall allow', 'Verify system allows the specified operation'),
        (r'Upon saving.*[Oo]fficial [Rr]eceipt', 'Verify system generates Official Receipt upon saving transaction'),
        (r'Upon saving.*[Oo]fficial [Rr]eceipt.*[Aa]gent', 'Verify system generates Official Receipt for agent transaction upon saving'),
        (r'Upon saving', 'Verify system processes and saves transaction correctly'),
        (r'Policy & LOB Retrieval.*no default', 'Verify system handles policy and LOB retrieval when no default value exists'),
        (r'Policy & LOB Retrieval', 'Verify system retrieves policy and line of business information'),
        (r'The system shall generate.*[Oo]fficial [Rr]eceipt', 'Verify system generates Official Receipt for the transaction'),
        (r'The system shall generate', 'Verify system generates the required output'),
        (r'The system shall filter', 'Verify system filters data according to business rules'),
        (r'The system shall validate', 'Verify system validates input according to business rules'),
        (r'The system shall process', 'Verify system processes transaction correctly'),
        (r'The system shall', 'Verify system performs the required operation'),
        (r'Renewal Premium.*PDF.*[Zz]ero [Aa]mount', 'Verify system validates zero amount for Renewal Premium with PDF bill type'),
        (r'Renewal Premium.*[Dd]irect [Bb]ill.*[Zz]ero [Aa]mount', 'Verify system validates zero amount for Renewal Premium with Direct Bill'),
        (r'Renewal Premium.*OLI.*[Zz]ero [Aa]mount', 'Verify system validates zero amount for Renewal Premium with OLI'),
        (r'Renewal Premium.*OLI.*45 day', 'Verify system processes Renewal Premium with OLI within 45-day grace period'),
        (r'Renewal Premium.*OLI', 'Verify system processes Renewal Premium with OLI correctly'),
        (r'Renewal Premium.*[Mm]ode [Cc]hange.*[Zz]ero', 'Verify system validates zero amount for Renewal Premium with mode change'),
        (r'Renewal Premium.*[Mm]ode [Cc]hange', 'Verify system processes Renewal Premium with mode change'),
        (r'Renewal Premium', 'Verify system processes Renewal Premium correctly'),
    ]

    for pattern, replacement in patterns:
        if re.search(pattern, title, re.IGNORECASE):
            return replacement

    # Generic fallback: convert "The system shall X" -> "Verify system X"
    m = re.match(r'The system shall\s+(.+)', title, re.IGNORECASE)
    if m:
        return f'Verify system {m.group(1).rstrip(".")}'

    # Last resort: prepend Verify
    return f'Verify {title}'


def generate_steps(scenario, precond, expected, test_data):
    """Generate test steps, each paired with its own expected result."""
    s = scenario.lower()
    td = test_data.lower()
    pre = precond.lower()

    # Each entry: (action, expected_result)
    pairs = []

    # Step 1: login
    pairs.append((
        'Log in to CDS Front Collection as Teller.',
        'User is successfully logged in to CDS Front Collection as Teller.'
    ))

    # Step 2: navigate
    if 'agent' in s or 'agent tran' in s:
        nav = 'Navigate to the Agent Transaction screen.'
        nav_exp = 'Agent Transaction screen is displayed.'
    elif 'group' in s or 'gpas' in s:
        nav = 'Navigate to the Group Policy Collection screen.'
        nav_exp = 'Group Policy Collection screen is displayed.'
    elif re.search(r'\bssi\b', s):
        nav = 'Navigate to the SSI Collection screen.'
        nav_exp = 'SSI Collection screen is displayed.'
    elif 'non-policy' in s or 'non policy' in s:
        nav = 'Navigate to the Non-Policy Transaction screen.'
        nav_exp = 'Non-Policy Transaction screen is displayed.'
    elif 'eod' in s or 'end of day' in s:
        nav = 'Navigate to the EOD Procedures screen.'
        nav_exp = 'EOD Procedures screen is displayed.'
    elif 'gl account' in s or 'gl mapping' in s:
        nav = 'Navigate to the GL Account Mapping screen.'
        nav_exp = 'GL Account Mapping screen is displayed.'
    elif 'collection header' in s or 'header setup' in s:
        nav = 'Navigate to the Collection Header Setup screen.'
        nav_exp = 'Collection Header Setup screen is displayed.'
    else:
        nav = 'Navigate to the Collection transaction screen.'
        nav_exp = 'Collection transaction screen is displayed.'
    pairs.append((nav, nav_exp))

    # Step 3: data entry / lookup
    if 'agent' in s or 'agent tran' in s:
        pairs.append((
            'Enter the agent code and retrieve agent details.',
            'Agent details are retrieved and displayed correctly.'
        ))
    elif 'non-policy' in s or 'non policy' in s:
        pairs.append((
            'Select the applicable non-policy transaction type.',
            'Non-policy transaction type is selected and form is ready for input.'
        ))
    elif 'eod' in s or 'end of day' in s:
        pairs.append((
            'Review the pending transactions and initiate the EOD procedure.',
            'Pending transactions are listed and EOD procedure is initiated.'
        ))
    elif 'gl account' in s or 'gl mapping' in s:
        pairs.append((
            'Select the transaction type to view the GL account mapping.',
            'GL account mapping for the selected transaction type is displayed.'
        ))
    elif 'collection header' in s or 'header setup' in s:
        pairs.append((
            'Open the Collection Header and review the transaction fields displayed.',
            'Collection Header fields are displayed in the correct sequence.'
        ))
    elif 'policy' in s or 'policy' in td or 'policy' in pre:
        pn_match = re.search(r'Policy(?:\s*No)?[:\s]+([A-Z0-9]{6,})', test_data, re.IGNORECASE)
        pn = pn_match.group(1) if pn_match else 'A12345678'
        if 'invalid' in s or 'invalid' in pre or 'unknown prefix' in td:
            pairs.append((
                'Enter an invalid policy number (e.g., X99999999) in the Policy Number field.',
                'System displays an error message for the invalid policy number.'
            ))
        elif 'not found in ingenium' in s or 'not found in ingenium' in pre:
            pairs.append((
                f'Enter a policy number that does not exist in Ingenium (e.g., {pn}).',
                'System does not find the policy in Ingenium and proceeds to search SUKI.'
            ))
        elif 'not found in suki' in s or 'not found in suki' in pre:
            pairs.append((
                f'Enter a policy number not found in Ingenium or SUKI (e.g., {pn}).',
                'System does not find the policy in SUKI and proceeds to search AUS.'
            ))
        elif 'not found in aus' in s or 'not found in aus' in pre:
            pairs.append((
                f'Enter a policy number not found in Ingenium, SUKI, or AUS (e.g., {pn}).',
                'System displays an error message indicating the policy was not found in any source.'
            ))
        elif 'group' in s or 'gpas' in s:
            pairs.append((
                'Enter a valid group policy number in the Policy Number field.',
                'System retrieves and displays the group policy details from GPAS.'
            ))
        else:
            pairs.append((
                f'Enter a valid policy number (e.g., {pn}) and press Enter or click Search.',
                f'System retrieves and displays the policy information for {pn}.'
            ))
    else:
        pairs.append((
            'Enter the required transaction details.',
            'Transaction details are accepted and displayed correctly.'
        ))

    # Step 4: action / amount / selection
    if 'zero amount' in s or 'zero amount' in pre or '₱0.00' in test_data or 'amount: ₱0' in td:
        pairs.append((
            'Enter ₱0.00 as the transaction amount.',
            'System flags the zero amount entry for validation.'
        ))
    elif 'below minimum' in s or 'below minimum' in pre or '₱400' in test_data:
        amt_match = re.search(r'₱[\d,]+', test_data)
        amt = amt_match.group(0) if amt_match else '₱400'
        pairs.append((
            f'Enter {amt} as the transaction amount (below the ₱500 minimum).',
            f'System flags {amt} as below the minimum allowable amount.'
        ))
    elif 'minimum amount' in s or 'at minimum' in s or '₱500 (minimum)' in td:
        pairs.append((
            'Enter ₱500.00 as the transaction amount (at the minimum threshold).',
            'System accepts ₱500.00 as a valid minimum amount.'
        ))
    elif 'trancode' in s or 'tran code' in s or 'trancode' in pre:
        pairs.append((
            'Observe the TranCode dropdown and select the applicable TranCode for the transaction.',
            'Only applicable TranCodes are shown in the dropdown based on the policy type.'
        ))
    elif 'lob' in s or 'line of business' in s:
        pairs.append((
            'Observe the Line of Business field auto-populated by the system based on the policy prefix.',
            'Line of Business field is correctly auto-populated based on the policy prefix.'
        ))
    elif 'due date' in s or 'grace period' in s:
        td_match = re.search(r'Payment Date[:\s]+(.+?)(?:\||$)', test_data, re.IGNORECASE)
        pd_val = td_match.group(1).strip() if td_match else 'as per test data'
        pairs.append((
            f'Set the payment date to {pd_val}.',
            f'Payment date is set to {pd_val} and due date validation is applied.'
        ))
    elif 'status' in s and ('active' in s or 'pending' in s or 'lapsed' in s):
        status_match = re.search(r'Policy with (\w+) status', test_data, re.IGNORECASE)
        status = status_match.group(1) if status_match else 'the specified'
        pairs.append((
            f'Confirm the policy status is {status} and proceed with the transaction.',
            f'Policy status is confirmed as {status}.'
        ))
    elif 'official receipt' in s or 'or generation' in s:
        pairs.append((
            'Enter the required transaction details and click Post/Save.',
            'Transaction details are accepted and the system processes the request.'
        ))
    elif 'loan' in s:
        pairs.append((
            'Select the applicable loan TranCode and enter the loan payment amount.',
            'Loan TranCode is selected and loan payment amount is accepted.'
        ))
    elif 'dividend' in s:
        pairs.append((
            'Select the dividend TranCode and enter the applicable dividend amount.',
            'Dividend TranCode is selected and dividend amount is accepted.'
        ))
    elif 'top-up' in s or 'top up' in s:
        pairs.append((
            'Select the Top-Up TranCode and enter the top-up amount.',
            'Top-Up TranCode is selected and top-up amount is accepted.'
        ))
    elif 'mode change' in s:
        pairs.append((
            'Select the new payment mode and enter the applicable premium amount.',
            'New payment mode is selected and premium amount is updated accordingly.'
        ))
    elif 'pdf' in s or 'premium deposit fund' in s:
        pairs.append((
            'Select the PDF TranCode and enter the premium deposit fund amount.',
            'PDF TranCode is selected and premium deposit fund amount is accepted.'
        ))
    elif 'oli' in s or 'old loan interest' in s:
        pairs.append((
            'Verify the OLI amount is computed and displayed; enter the total amount including OLI.',
            'OLI amount is correctly computed and the total amount including OLI is accepted.'
        ))
    elif 'epon' in s:
        pairs.append((
            'Select the EPON TranCode and enter the applicable amount.',
            'EPON TranCode is selected and the applicable amount is accepted.'
        ))
    elif 'supervisor' in s or 'override' in s:
        pairs.append((
            'Attempt the transaction that requires supervisor override.',
            'System prompts for supervisor credentials to authorize the override.'
        ))
    elif 'representative' in s:
        pairs.append((
            'Open the Representative modal and verify the displayed information.',
            'Representative modal displays the correct agent/representative information.'
        ))
    elif 'invoice' in s:
        pairs.append((
            'Trigger invoice generation for the transaction.',
            'Invoice is generated with the correct transaction details.'
        ))
    elif 'eod' in s or 'end of day' in s:
        pairs.append((
            'Confirm and execute the EOD procedure.',
            'EOD procedure is executed and summary report is generated.'
        ))
    elif 'payment method' in s:
        pairs.append((
            'Select the applicable payment method (Cash/Check/Online).',
            'Selected payment method is accepted and applied to the transaction.'
        ))
    elif 'year' in s and 'determination' in s:
        pairs.append((
            'Observe the year determination logic applied by the system.',
            'System correctly determines and applies the applicable policy year.'
        ))
    elif 'modal premium' in s:
        pairs.append((
            'Verify the modal premium displayed matches the policy data.',
            'Modal premium displayed matches the expected value from the policy data.'
        ))
    elif 'pdf cap' in s:
        pairs.append((
            'Enter an amount that exceeds the PDF cap limit.',
            'System flags the amount as exceeding the PDF cap limit.'
        ))
    elif 'customer' in s:
        pairs.append((
            'Enter the customer details and perform the required operation.',
            'Customer details are accepted and the operation is performed successfully.'
        ))
    elif 'amount' in s or 'premium' in s:
        amt_match = re.search(r'₱[\d,]+(?:\.\d+)?', test_data)
        amt = amt_match.group(0) if amt_match else 'the applicable premium amount'
        pairs.append((
            f'Enter {amt} as the transaction amount.',
            f'System accepts {amt} as the transaction amount.'
        ))
    elif 'collection header' in s or 'header setup' in s:
        pairs.append((
            'Verify all required fields are displayed in the correct sequence.',
            'All required fields are present and displayed in the correct sequence.'
        ))
    elif 'gl account' in s:
        pairs.append((
            'Verify the GL account mapping displayed for the transaction type.',
            'GL account mapping is correctly displayed per the business rules.'
        ))
    else:
        pairs.append((
            'Enter the required transaction details.',
            'Transaction details are accepted and displayed correctly.'
        ))

    # Step 5+: post/save and final verification
    if 'official receipt' in s or 'or generation' in s:
        pairs.append((
            'Verify the Official Receipt (OR) is generated with the correct details.',
            'Official Receipt is generated with the correct transaction details, amount, and policy information.'
        ))
    elif 'zero amount' in s or '₱0.00' in test_data:
        pairs.append((
            'Attempt to post the transaction.',
            'System prevents posting and displays a validation error for zero amount.'
        ))
        pairs.append((
            'Verify the system displays the appropriate validation message.',
            'Validation message is displayed indicating zero amount is not allowed.'
        ))
    elif 'below minimum' in s or 'below minimum' in pre:
        pairs.append((
            'Attempt to post the transaction.',
            'System rejects the transaction due to amount being below the minimum.'
        ))
        pairs.append((
            'Verify the system rejects the transaction and displays an error message.',
            'Error message is displayed indicating the amount is below the minimum allowable value.'
        ))
    elif 'invalid' in s or 'not found' in s:
        pairs.append((
            'Press Enter or attempt to proceed.',
            'System processes the input and applies the appropriate validation.'
        ))
        pairs.append((
            'Verify the system displays the appropriate error or fallback behavior.',
            'System displays the correct error message or fallback behavior as per business rules.'
        ))
    elif 'supervisor' in s or 'override' in s:
        pairs.append((
            'Have the supervisor enter their credentials to authorize the override.',
            'Supervisor credentials are accepted and the override is authorized.'
        ))
        pairs.append((
            'Confirm the transaction is processed after supervisor approval.',
            'Transaction is successfully processed after supervisor approval.'
        ))
    elif 'eod' in s or 'end of day' in s:
        pairs.append((
            'Verify the EOD summary and reports are generated correctly.',
            'EOD summary and all required reports are generated with accurate data.'
        ))
    elif 'lob' in s or 'line of business' in s:
        pairs.append((
            'Verify the Line of Business value auto-populated correctly based on the policy prefix.',
            'Line of Business value matches the expected value for the given policy prefix.'
        ))
    elif 'due date' in s or 'grace period' in s:
        pairs.append((
            'Enter the applicable premium amount and click Post/Save.',
            'Premium amount is accepted and transaction is submitted for processing.'
        ))
        pairs.append((
            'Verify the transaction is processed with the correct GL accounts.',
            'Transaction is processed successfully and correct GL accounts are applied.'
        ))
    elif 'status' in s:
        pairs.append((
            'Attempt to post the transaction.',
            'System evaluates the policy status and applies the appropriate business rules.'
        ))
        pairs.append((
            'Verify the system handles the policy status appropriately.',
            'System handles the policy status correctly as per the defined business rules.'
        ))
    elif 'trancode' in s:
        pairs.append((
            'Verify only the applicable TranCodes are shown in the dropdown based on policy type.',
            'Dropdown shows only the TranCodes applicable to the current policy type.'
        ))
    elif 'display' in s or 'retriev' in s:
        pairs.append((
            'Verify the policy information is displayed correctly on screen.',
            'All policy information fields are displayed accurately on screen.'
        ))
    elif 'collection header' in s or 'header setup' in s:
        pairs.append((
            'Verify the field validations and sequencing work as expected.',
            'Field validations trigger correctly and fields are sequenced as per business rules.'
        ))
    elif 'gl account' in s:
        pairs.append((
            'Verify the GL accounts are correctly mapped per the business rules.',
            'GL accounts are correctly mapped and match the expected configuration.'
        ))
    else:
        pairs.append((
            'Click Post/Save to submit the transaction.',
            'Transaction is submitted and system begins processing.'
        ))
        pairs.append((
            'Verify the system processes the transaction and displays the confirmation.',
            'Transaction is processed successfully and a confirmation message is displayed.'
        ))

    steps_lines = [f'{i}. {action}' for i, (action, _) in enumerate(pairs, 1)]
    expected_lines = [f'{i}. {exp}' for i, (_, exp) in enumerate(pairs, 1)]
    return '\n'.join(steps_lines), '\n'.join(expected_lines)



def parse_module_files(base_dir='test_cases_by_module'):
    """Parse all module .md files and return structured rows with real data."""
    import glob, os

    # Folder -> display module name mapping
    folder_module_map = {
        'Agent Transactions': 'Agent Transactions',
        'Collection Reversal': 'Collection Reversal',
        'Group Policy': 'Group Policy',
        'Others': 'Others',
        'Policy Direct Marketing': 'Policy Direct Marketing',
        'Policy Regular (BASE)': 'Policy Regular (BASE)',
        'Policy VUL': 'Policy VUL',
        'Regular (Balance Payment)': 'Regular (Balance Payment)',
        'SSI': 'SSI',
    }

    def get_table_field(field_name, block):
        pattern = rf'\|\s*{re.escape(field_name)}\s*\|\s*(.+?)\s*\|'
        m = re.search(pattern, block, re.IGNORECASE)
        if not m:
            return ''
        val = m.group(1).strip()
        val = re.sub(r'<br>\s*-\s*', '\n- ', val)
        val = re.sub(r'<br>', '\n', val)
        val = re.sub(r'<[^>]+>', '', val)
        return val.strip()

    all_rows = []
    for filepath in sorted(glob.glob(f'{base_dir}/**/*.md', recursive=True)):
        if 'README' in filepath:
            continue
        folder = os.path.basename(os.path.dirname(filepath))
        module = folder_module_map.get(folder, folder)

        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()

        tc_parts = re.split(r'\n(?=## TC-)', content)
        for tc_block in tc_parts:
            tc_match = re.match(r'## (TC-[\w-]+):\s*(.+)', tc_block)
            if not tc_match:
                continue
            tc_id = tc_match.group(1).strip()

            user_story = get_table_field('User Story', tc_block)
            scenario   = get_table_field('Scenario', tc_block)
            precond    = get_table_field('Preconditions', tc_block)
            expected   = get_table_field('Expected Result', tc_block)
            test_data  = get_table_field('Test Data', tc_block)

            # Use Scenario as title; normalize to start with Verify
            raw_title = scenario if scenario else tc_match.group(2).strip()
            title = normalize_title(raw_title)

            # Clean up preconditions bullet formatting
            precond_lines = [l.strip().lstrip('-').strip() for l in precond.splitlines() if l.strip()]
            precond_clean = '\n'.join(f'- {l}' if not l.startswith('-') else l for l in precond_lines)

            # Clean up expected result bullet formatting
            exp_lines = [l.strip().lstrip('-').strip() for l in expected.splitlines() if l.strip()]
            exp_clean = '\n'.join(f'- {l}' if not l.startswith('-') else l for l in exp_lines)

            # Generate specific steps and per-step expected results
            steps, steps_expected = generate_steps(raw_title, precond, expected, test_data)

            all_rows.append({
                'Module': module,
                'Test Case ID': tc_id,
                'User Story': user_story,
                'Test Case Title': title,
                'Preconditions': precond_clean,
                'Steps': steps,
                'Expected Results': steps_expected,
                'Test Data': test_data,
                'Status': 'Not Started',
                '_tc_key': f'{module}|{tc_id}',
            })

    # Deduplicate within the same User Story — keep first occurrence of each (Module, UserStory, Title, Precond, Expected)
    # For rows with no User Story, fall back to TC ID as the unique key
    seen_keys = set()
    deduped = []
    for r in all_rows:
        us = r['User Story'] or None
        if us:
            key = (r['Module'], us, r['Test Case Title'], r['Preconditions'], r['Expected Results'])
        else:
            # No user story — deduplicate only by TC ID to avoid removing valid distinct TCs
            key = (r['Module'], r['Test Case ID'])
        if key not in seen_keys:
            seen_keys.add(key)
            deduped.append(r)
    removed = len(all_rows) - len(deduped)
    if removed:
        print(f'  [dedup] Removed {removed} duplicate test cases within same user story')
    return deduped


def fast_merge(ws, min_row, min_col, max_row, max_col):
    """Bypass openpyxl's O(n) overlap check for a significant speed boost."""
    from openpyxl.worksheet.cell_range import CellRange
    cr = CellRange(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col)
    ws.merged_cells.ranges.add(cr)
    # Clear interior cells so Excel doesn't complain
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            if row == min_row and col == min_col:
                continue
            ws._cells.pop((row, col), None)


def create_excel_from_module_files(rows, output_path):
    """Creates Excel with one sheet per module. Each expected result bullet = one row."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    seen_modules = []
    for r in rows:
        if r['Module'] not in seen_modules:
            seen_modules.append(r['Module'])

    headers = ['Module', 'Test Case ID', 'User Story', 'Test Case Title', 'Preconditions', 'Steps', 'Expected Results', 'Test Data', 'Status']
    col_widths = [22, 14, 42, 45, 32, 45, 45, 35, 14]

    # Column indices (1-based) for columns that get merged per TC
    merge_col_indices = [headers.index(h) + 1 for h in ['Module', 'Test Case ID', 'User Story', 'Test Case Title', 'Preconditions', 'Steps', 'Test Data']]

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap_top = Alignment(vertical='top', wrap_text=True)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_even = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    fill_odd  = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    status_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    for module in seen_modules:
        mod_rows = [r for r in rows if r['Module'] == module]
        sheet_name = re.sub(r'[\\/*?:\[\]]', '', module)[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Header row
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        ws.row_dimensions[1].height = 30

        tc_counter = 0
        tc_color = {}
        excel_row = 2

        us_col_idx = headers.index('User Story') + 1
        exp_col_idx = headers.index('Expected Results') + 1
        # Columns to merge per TC (all except User Story and Expected Results)
        tc_merge_cols = [c for c in merge_col_indices if c != us_col_idx]

        # Pre-compute row spans — one row per expected result line per TC
        row_plan = []  # (tc_key, us_text, start_row, end_row, n_exp_lines)
        temp_row = 2
        for row in mod_rows:
            tc_key = row['_tc_key']
            exp_lines = [l.strip() for l in row['Expected Results'].splitlines() if l.strip()]
            if not exp_lines:
                exp_lines = ['']
            n = len(exp_lines)
            row_plan.append((tc_key, row.get('User Story', ''), temp_row, temp_row + n - 1, exp_lines))
            temp_row += n

        # Write all cells
        for (tc_key, us_text, s, e, exp_lines), row in zip(row_plan, mod_rows):
            if tc_key not in tc_color:
                tc_color[tc_key] = tc_counter % 2
                tc_counter += 1
            fill = fill_even if tc_color[tc_key] == 0 else fill_odd

            for idx, exp_line in enumerate(exp_lines):
                for col, key in enumerate(headers, 1):
                    if key == 'Expected Results':
                        val = exp_line
                    elif key == 'Status':
                        val = 'Not Started'
                    elif idx == 0:
                        val = row.get(key, '')
                    else:
                        val = ''  # will be merged

                    cell = ws.cell(row=excel_row, column=col, value=val)
                    cell.border = border
                    cell.fill = fill
                    cell.alignment = wrap_top

                    if key == 'Status':
                        cell.fill = status_fill
                        cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

                ws.row_dimensions[excel_row].height = 20
                excel_row += 1

        # Merge TC-level columns across all expected result rows of each TC
        for (tc_key, us_text, s, e, _) in row_plan:
            if e > s:
                for col_idx in tc_merge_cols:
                    fast_merge(ws, s, col_idx, e, col_idx)
                    ws.cell(row=s, column=col_idx).alignment = Alignment(vertical='top', wrap_text=True)

        # Merge User Story column for contiguous blocks of the same user story
        if row_plan:
            cur_us = row_plan[0][1]
            cur_start = row_plan[0][2]
            cur_end = row_plan[0][3]
            for (_, us_text, s, e, _) in row_plan[1:]:
                if us_text == cur_us:
                    cur_end = e
                else:
                    if cur_end > cur_start:
                        fast_merge(ws, cur_start, us_col_idx, cur_end, us_col_idx)
                        ws.cell(row=cur_start, column=us_col_idx).alignment = Alignment(vertical='top', wrap_text=True)
                    cur_us = us_text
                    cur_start = s
                    cur_end = e
            if cur_end > cur_start:
                fast_merge(ws, cur_start, us_col_idx, cur_end, us_col_idx)
                ws.cell(row=cur_start, column=us_col_idx).alignment = Alignment(vertical='top', wrap_text=True)

        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'
        print(f'  Sheet "{sheet_name}": {len(mod_rows)} test cases')

    wb.save(output_path)
    print(f'Saved: {output_path} ({len(rows)} total test cases)')


if __name__ == '__main__':
    rows = parse_module_files('test_cases_by_module')
    create_excel_from_module_files(rows, 'test-cases-all-modules.xlsx')
